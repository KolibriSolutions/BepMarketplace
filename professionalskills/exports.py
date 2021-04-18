#  Bep Marketplace ELE
#  Copyright (c) 2016-2021 Kolibri Solutions
#  License: See LICENSE file or https://github.com/KolibriSolutions/BepMarketplace/blob/master/LICENSE

from django.conf import settings
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from general_view import timestamp
from professionalskills.models import FileType, StaffResponse
from results.models import GradeCategory
from timeline.utils import get_timeslot
from .models import StaffResponseFileAspectResult
from django.utils import timezone


def get_prv_type_xlsx(prv, des):
    """

    :param prv:
    :return:
    """
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    ws.title = prv.Name

    ws['A1'] = f'Students grades from {settings.NAME_PRETTY} of {prv.Name} in {prv.TimeSlot.Name}'
    ws['A1'].style = 'Headline 2'
    ws['F1'] = "Exported on: " + timestamp()

    ws['A2'] = 'Name'
    ws['B2'] = 'Final Bachelor Project'

    ws['A3'] = 'Time slot'
    ws['B3'] = prv.TimeSlot.Name

    ws['A4'] = 'Test'
    ws['B4'] = prv.Name

    ws['A5'] = 'Deadline'
    ws['B5'] = prv.Deadline.strftime("%A %d %B %Y")

    ws['A6'] = 'Opportunity'
    ws['B6'] = ''

    ws['A7'] = ''
    ws['B7'] = ''

    header = ["Student id", "Student name", "Hand-in", "Grade", "Comment", ]
    aspects = prv.aspects.all().order_by('id')
    for a in aspects:
        header.append(a.Name)

    ws.column_dimensions['B'].width = 25  # name
    # ws.column_dimensions['C'].width = 25  # Project
    # ws.column_dimensions['D'].width = 25  # responsible
    ws.column_dimensions['E'].width = 25  # comment

    ws.append(header)
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
            'V', 'W', 'X', 'Y', 'Z'][:len(header)]
    for col in cols:
        ws[col + '8'].style = 'Headline 3'

    for d in des:
        row = [d.Student.usermeta.Studentnumber, d.Student.usermeta.Fullname]
        files = d.files.filter(Type=prv).order_by('-id')
        if files.exists():
            file = files.first()
            row.append(timezone.localtime(file.TimeStamp).strftime("%Y-%m-%d %H:%M"))
            try:
                row.append(StaffResponse.StatusOptionsOsiris[file.staffresponse.Status])
                row.append(file.staffresponse.Explanation)
                for aspect in aspects:
                    try:
                        ar = file.staffresponse.aspects.get(Aspect=aspect)
                        row.append(ar.get_Grade_display())
                    except StaffResponseFileAspectResult.DoesNotExist:
                        row.append('-')
            except StaffResponse.DoesNotExist:
                row.append('-')
            if files.count() > 1:
                row.append('Warning, multiple files, last one shown.')
        else:  # no file
            row.append('No file')
        ws.append(row)
    return save_virtual_workbook(wb)
