#  Bep Marketplace ELE
#  Copyright (c) 2016-2020 Kolibri Solutions
#  License: See LICENSE file or https://github.com/KolibriSolutions/BepMarketplace/blob/master/LICENSE
#
from datetime import datetime

from django.conf import settings
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from general_view import timestamp
from professionalskills.models import FileType, StaffResponse
from proposals.utils import get_share_link
from results.models import GradeCategory
from timeline.utils import get_timeslot


def get_list_students_xlsx(des, typ):
    """
    Export students and their grades.
    List all students with their proposal and grades.

    :param des:
    :param typ:
    :return:
    """

    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    ws.title = "grades"

    ws['A1'] = "Students grades from {}".format(settings.NAME_PRETTY)
    ws['A1'].style = 'Headline 2'
    ws['F1'] = "Exported on: " + timestamp()

    header = ["Student id", "Student name", "Project", "Responsible teacher", "Assistants", "ECTS", "Track"]
    for t in typ:
        header.append(t.Name + " (" + str(t.Weight) + "%)")

    header.append("Total")
    header.append("Total rounded")

    prvs = list(FileType.objects.filter(TimeSlot=get_timeslot()))
    for prv in prvs:
        header.append(prv.Name)

    ws.column_dimensions['B'].width = 25  # name
    ws.column_dimensions['C'].width = 25  # Project
    ws.column_dimensions['D'].width = 25  # responsible
    ws.column_dimensions['E'].width = 25  # assistants

    ws.append(header)
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
            'V', 'W', 'X', 'Y', 'Z'][:len(header)]
    for col in cols:
        ws[col + '2'].style = 'Headline 3'

    cats = GradeCategory.objects.filter(TimeSlot=get_timeslot())

    for d in des:
        reslist = []
        for c in cats:
            try:
                reslist.append(d.results.get(Category=c).Grade)
            except:
                reslist.append('-')
        row = [d.Student.usermeta.Studentnumber, d.Student.usermeta.get_nice_fullname(), d.Proposal.Title,
               d.Proposal.ResponsibleStaff.usermeta.get_nice_name()]
        assistants = ''
        for a in d.Proposal.Assistants.all():
            assistants += a.usermeta.get_nice_name() + "; "
        row.append(assistants)
        if d.Student.usermeta.EnrolledExt:
            row.append(15)
        else:
            row.append(10)
        row.append(d.Proposal.Track.Name)
        for r in reslist:
            row.append(r)
        row.append(round(d.TotalGrade(), 2))
        row.append(d.TotalGradeRounded())
        for prv in prvs:
            try:
                row.append(d.files.filter(Type=prv).order_by('-id')[0].staffresponse.Status)
            except IndexError:
                row.append('no file')
            except StaffResponse.DoesNotExist:
                row.append('no grading')

        ws.append(row)

    # second tab with prv data

    ws = wb.create_sheet(title='prv-aspects')

    ws['A1'] = "Students grades from {} (only Aspects related to PRV)".format(settings.NAME_PRETTY)
    ws['A1'].style = 'Headline 2'
    ws['F1'] = "Exported on: " + timestamp()
    header = ["Student id", "Student name", "Project", "Responsible teacher"]

    aspects = []  # list of PRV related aspects.
    for cat in cats:  # each category
        for aspect in cat.aspects.all():
            if 'prv' in aspect.Name.lower() or 'prv' in aspect.Description.lower():
                aspects.append([cat, aspect])
                header.append(aspect.Name)

    ws.append(header)
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
            'V', 'W', 'X', 'Y', 'Z'][:len(header)]
    for col in cols:
        ws[col + '2'].style = 'Headline 3'

    for colnr in range(1, len(header)):  # set width
        ws.column_dimensions[cols[colnr]].width = 25

    for d in des:
        reslist = []
        for c, a in aspects:
            try:
                reslist.append(d.results.get(Category=c).aspectresults.get(CategoryAspect=a).Grade)
            except:
                reslist.append('-')
        row = [d.Student.usermeta.Studentnumber, d.Student.usermeta.get_nice_fullname(), d.Proposal.Title,
               d.Proposal.ResponsibleStaff.usermeta.get_nice_name()]
        for r in reslist:
            row.append(r)
        ws.append(row)

    return save_virtual_workbook(wb)


def get_list_distributions_xlsx(proposals):
    """
    Excel export of proposals with distributions.
    Lists all proposals with their student.

    :param proposals:
    :return:
    """
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    ws.title = "distributions"

    ws['A1'] = "Proposals with students and staff from {}".format(settings.NAME_PRETTY)
    ws['A1'].style = 'Headline 2'
    ws['F1'] = "Exported on: " + str(datetime.now())

    header = ["Title", "Track", "Research group", "Responsible", "Assistants", "Chosen", "Student Emails",
              "Staff Emails"]

    ws.column_dimensions['A'].width = 25  # title
    ws.column_dimensions['D'].width = 25  # responsible
    ws.column_dimensions['E'].width = 25  # assistants
    ws.column_dimensions['F'].width = 25  # chosen
    ws.column_dimensions['G'].width = 25  # student mail
    ws.column_dimensions['H'].width = 25  # staff mail

    ws.append(header)
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[col + '2'].style = 'Headline 3'

    for p in proposals:
        des = p.distributions.filter(TimeSlot=get_timeslot())
        row = [p.Title, p.Track.__str__(), p.Group.__str__(),
               p.ResponsibleStaff.usermeta.get_nice_name()]
        assistants = ''
        for a in p.Assistants.all():
            assistants += a.usermeta.get_nice_name() + "; "
        row.append(assistants)
        stds = ''
        stdsmail = ''
        for d in des:
            stdsmail += d.Student.email + '; '
            try:
                stds += d.Student.usermeta.get_nice_fullname() + " (" + d.Student.usermeta.Studentnumber + "); "
            except:
                stds += d.Student.usermeta.get_nice_fullname() + "; "
        row.append(stds)
        row.append(stdsmail)
        emails = '{};'.format(p.ResponsibleStaff.email)
        for a in p.Assistants.all():
            emails += '{};'.format(a.email)
        row.append(emails)
        # row[4].alignment = Alignment(wrapText=True)
        # row[5].alignment = Alignment(wrapText=True)
        ws.append(row)
    return save_virtual_workbook(wb)


def get_list_projects_xlsx(proposals):
    """
    Excel export of proposals with sharelinks

    :param proposals:
    :return:
    """

    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    ws.title = "BEP Projects"

    ws['A1'] = 'Projects from {}'.format(settings.NAME_PRETTY)
    ws['A1'].style = 'Headline 2'
    ws['F1'] = "Exported on: " + str(datetime.now())

    header = ['Title', 'Track', 'Research Group', 'Responsible', 'email', 'Sharelink']
    h = ['A', 'B', 'C', 'D', 'E', 'F']

    ws.append(header)

    for hr in h:
        ws.column_dimensions[hr].width = 25
        ws[hr + '2'].style = 'Headline 3'
    ws.column_dimensions['F'].width = 100

    for p in proposals:
        ws.append([p.Title, str(p.Track), str(p.Group), p.ResponsibleStaff.usermeta.get_nice_name(), p.ResponsibleStaff.email,
                   get_share_link(p.pk)])

    return save_virtual_workbook(wb)
