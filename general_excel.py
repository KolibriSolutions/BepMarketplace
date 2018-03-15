"""
Excel export functions, for various excel lists. Called from specific views.
"""
from datetime import datetime

from django.db.models import Count, Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from results.models import GradeCategory
from timeline.utils import get_timeslot


def timestamp():
    """
    Timestamp for a xls export

    :return:
    """
    return "{:%Y-%m-%d %H:%M:%S}".format(datetime.now())


def listStudentsXls(des, typ):
    """
    Export students and their grades.
    List all students with their proposal and grades.
    
    :param des: 
    :param typ: 
    :return: 
    """

    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    ws.title = "grades"

    ws['A1'] = "Students grades from Bep Marketplace"
    ws['A1'].style = 'Headline 2'
    ws['F1'] = "Exported on: " + timestamp()

    header = ["Student id", "Student name", "Project", "Responsible teacher", "Assistants", "ECTS", "Track"]
    for t in typ:
        header.append(t.Name + " (" +  str(t.Weight) + "%)")

    header.append("Total")
    header.append("Total rounded")

    ws.column_dimensions['B'].width = 25  # name
    ws.column_dimensions['C'].width = 25  # Project
    ws.column_dimensions['D'].width = 25  # responsible
    ws.column_dimensions['E'].width = 25  # assistants

    ws.append(header)
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z'][:len(header)]
    for col in cols:
        ws[col + '2'].style = 'Headline 3'

    cats = GradeCategory.objects.filter(TimeSlot=get_timeslot())
    for d in des:
        reslist = []
        for c in cats:
            try:
                reslist.append(d.results.get(Category=c).Grade)
            except:
                reslist.append('-')
        row = [d.Student.usermeta.Studentnumber, d.Student.usermeta.Fullname, d.Proposal.Title,
               d.Proposal.ResponsibleStaff.first_name + " " + d.Proposal.ResponsibleStaff.last_name]
        assistants = ''
        for a in d.Proposal.Assistants.all():
            assistants += a.first_name + " " + a.last_name + "; "
        row.append(assistants)
        row.append(d.Proposal.ECTS)
        row.append(d.Proposal.Track.Name)
        for r in reslist:
            row.append(r)
        row.append(round(d.TotalGrade(), 2))
        row.append(d.TotalGradeRounded())
        ws.append(row)

    return save_virtual_workbook(wb)


def listDistributionsXls(proposals):
    """
    Excel export of proposals with distributions.
    Lists all proposals with their student.
    
    :param proposals: 
    :return: 
    """
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    ws.title = "distributions"

    ws['A1'] = "Proposals with students and staff from Bep Marketplace"
    ws['A1'].style = 'Headline 2'
    ws['F1'] = "Exported on: " + str(datetime.now())

    header = ["Title", "Track", "Research group","Responsible","Assistants","Chosen", "Student Emails", "Staff Emails"]

    ws.column_dimensions['A'].width = 25  # title
    ws.column_dimensions['D'].width = 25  # responsible
    ws.column_dimensions['E'].width = 25  # assistants
    ws.column_dimensions['F'].width = 25  # chosen
    ws.column_dimensions['G'].width = 25  # student mail
    ws.column_dimensions['H'].width = 25  # staff mail

    ws.append(header)
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[col + '2'].style = 'Headline 3'

    for p in proposals:
        des = p.distributions.filter(Timeslot=get_timeslot())
        row = [p.Title, p.Track.__str__(), p.Group.__str__(),
               p.ResponsibleStaff.first_name + ' ' + p.ResponsibleStaff.last_name]
        assistants = ''
        for a in p.Assistants.all():
            assistants += a.first_name + " " + a.last_name + "; "
        row.append(assistants)
        stds = ''
        stdsmail = ''
        for d in des:
            stdsmail += d.Student.email + '; '
            try:
                stds += d.Student.usermeta.Fullname + " (" + d.Student.usermeta.Studentnumber + "); "
            except:
                stds += d.Student.usermeta.Fullname + "; "
        row.append(stds)
        row.append(stdsmail)
        emails = '{};'.format(p.ResponsibleStaff.email)
        for a in p.Assistants.all():
            emails += '{};'.format(a.email)
        row.append(emails)
        #row[4].alignment = Alignment(wrapText=True)
        #row[5].alignment = Alignment(wrapText=True)
        ws.append(row)
    return save_virtual_workbook(wb)


def listStaffXls(staff):
    """
    Lists all staff from the marketplace, with number of proposals and distributions
    
    :param staff: 
    :return: 
    """
    def nint(nr):
        """

        :param nr:
        :return:
        """
        if nr is None:
            return 0
        else:
            return int(nr)

    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    ws.title = "staff"

    ws['A1'] = "Staff from Bep Marketplace"
    ws['A1'].style = 'Headline 2'
    ws['F1'] = "Exported on: " + timestamp()
    header = ["Name", "Email", "Proposals responsible", "Proposals assistant", "Proposals total", "Distribution responsible ","Distributions assistant", "Distributions total"]

    ws.column_dimensions['A'].width = 25  # name
    ws.column_dimensions['B'].width = 25  # mail

    ws.append(header)

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[col + '2'].style = 'Headline 3'

    for s in staff:
        pt1 = s.proposalsresponsible.count()
        pt2 = s.proposals.count()
        pts = pt1+pt2
        dt1 = nint(s.proposalsresponsible.all().annotate(Count('distributions')).aggregate(Sum('distributions__count'))[
                 'distributions__count__sum'])
        dt2 = nint(s.proposals.all().annotate(Count('distributions')).aggregate(Sum('distributions__count'))[
                 'distributions__count__sum'])
        dts = dt1+dt2
        row = [s.get_full_name(), s.email,
               pt1, pt2, pts,
               dt1, dt2, dts
            ]
        ws.append(row)
    return save_virtual_workbook(wb)


def listPresentationsXls(sets):
    """
    Lists the presentations planning with a set on each tab. Same formatting as before marketplace
    
    :param sets:
    :return:
    """
    codeExt = "5XED0"
    codeBEP = "5XEC0"
    wb = Workbook()
    wb.remove_sheet(wb.active)
    wss = []
    for pset in sets:
        tit=str(pset.id)+" ("+pset.Track.Name+")"
        ws= wb.create_sheet(title=tit)
        wss.append(ws)
        ws['C1'] = timezone.localtime(pset.DateTime).date().strftime("%A %d %B %Y")
        ws['C3'] = "Track: " + pset.Track.Name
        ws['C4'] = "Track head: " + pset.Track.Head.usermeta.Fullname
        ws['C5'] = "Exported on: " + timestamp()
        ws['C6'] = "Presentation room: " + pset.PresentationRoom.Name
        ws['C7'] = "Assessment room: " + pset.AssessmentRoom.Name
        assessors = ''
        for a in pset.Assessors.all():
            assessors += a.usermeta.Fullname + "; "
        ws['C8'] = 'Assessors: ' + assessors[:-2]
        ws['C9'] = ''

        ws['C1'].style = 'Headline 2'
        # courses span two columns
        ws['G9'] = 'Courses'
        ws.merge_cells('G9:H9')
        ws['G9'].style = 'Headline 3'
        ws['H9'].style = 'Headline 3'
        # custom dimensions
        ws.column_dimensions['C'].width = 25  # std name
        ws.column_dimensions['E'].width = 25  # responsible name
        ws.column_dimensions['F'].width = 25  # assistants
        ws.column_dimensions['G'].width = 6  # codeBEP
        ws.column_dimensions['H'].width = 6  # codeExt
        ws.column_dimensions['I'].width = 50  # project name

        header = ["Type", "Stud. id", "Name", "First name", "Responsible teacher", "Assistants",
                  codeBEP, codeExt, "Project", "Time", "Duration"]
        ws.append(header)
        for col in ['A','B','C','D','E','F','G','H','I','J','K']:
            ws[col+'10'].style = 'Headline 3'

        for slot in pset.timeslots.all():
            if slot.CustomType:
                row = [slot.get_CustomType_display(), '', '', '', '', '', '', '', '']
            else:
                d = slot.Distribution
                row = ['', d.Student.usermeta.Studentnumber, d.Student.usermeta.Fullname, d.Student.first_name,
                       d.Proposal.ResponsibleStaff.usermeta.Fullname]
                assistants = ''
                for a in d.Proposal.Assistants.all():
                    assistants += a.usermeta.Fullname + "; "
                row.append(assistants[:-2])
                row.append('x' if d.Student.usermeta.EnrolledBEP else '')
                row.append('x' if d.Student.usermeta.EnrolledExt else '')
                row.append(d.Proposal.Title)
            row.append(timezone.localtime(slot.DateTime).time().strftime("%H:%M"))
            row.append(slot.Duration())
            ws.append(row)
    return save_virtual_workbook(wb)