#  Bep Marketplace ELE
#  Copyright (c) 2016-2022 Kolibri Solutions
#  License: See LICENSE file or https://github.com/KolibriSolutions/BepMarketplace/blob/master/LICENSE

from django.conf import settings
from openpyxl import Workbook

from general_view import timestamp
from professionalskills.models import FileType, StaffResponse
from results.models import GradeCategory
from timeline.utils import get_timeslot


def get_list_students_xlsx(des, typ, timeslot):
    """
    Export students and their grades.
    List all students with their proposal and grades.

    :param des: list of distributions
    :param typ: list of grade types
    :param timeslot: timeslot object
    :return:
    """

    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    ws.title = "grades"

    ws['A1'] = f'Students grades from {settings.NAME_PRETTY} of {timeslot}'
    ws['A1'].style = 'Headline 2'
    ws['F1'] = "Exported on: " + timestamp()

    header = ["Student id", "Student name", "Project", "Responsible teacher", "Assistants", "Assessors", "Pres. Assessors", "ECTS", "Track"]
    for t in typ:
        header.append(t.Name + " (" + str(t.Weight) + "%)")

    header.append("Total")
    header.append("Total rounded")

    prvs = list(FileType.objects.filter(TimeSlot=timeslot))
    for prv in prvs:
        header.append(prv.Name)

    ws.column_dimensions['B'].width = 25  # name
    ws.column_dimensions['C'].width = 25  # Project
    ws.column_dimensions['D'].width = 25  # responsible
    ws.column_dimensions['E'].width = 25  # assistants
    ws.column_dimensions['F'].width = 25  # Assessors
    ws.column_dimensions['G'].width = 25  # Presentation Assessors

    ws.append(header)
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
            'V', 'W', 'X', 'Y', 'Z'][:len(header)]
    for col in cols:
        ws[col + '2'].style = 'Headline 3'

    cats = GradeCategory.objects.filter(TimeSlot=timeslot)

    for d in des:
        reslist = []
        for c in cats:
            try:
                reslist.append(d.results.get(Category=c).Grade)
            except:
                reslist.append('-')
        row = [d.Student.usermeta.Studentnumber, d.Student.usermeta.get_nice_fullname(), d.Proposal.Title,
               d.Proposal.ResponsibleStaff.usermeta.get_nice_name()]
        assistants = ''
        for a in d.Proposal.Assistants.all():
            assistants += a.usermeta.get_nice_name() + "; "
        row.append(assistants)
        if hasattr(d, 'presentationtimeslot'):
            ass1 = ''
            for a in d.presentationtimeslot.Presentations.Assessors.all():
                ass1 += a.usermeta.get_nice_name() + "; "
            row.append(ass1)
            ass2 = ''
            for a in d.presentationtimeslot.Presentations.PresentationAssessors.all():
                ass2 += a.usermeta.get_nice_name() + "; "
            row.append(ass2)
        else:
            row.append('-')
            row.append('-')
        if d.Student.usermeta.EnrolledExt:
            row.append(15)
        else:
            row.append(10)
        row.append(d.Proposal.Track.Name)
        for r in reslist:
            row.append(r)
        row.append(round(d.TotalGrade(), 2))
        row.append(d.TotalGradeRounded())
        for prv in prvs:
            files = d.files.filter(Type=prv).order_by('-id')
            if files.exists():
                cell = ''
                for file in files:
                    try:
                        cell += file.staffresponse.Status + '; '
                    except StaffResponse.DoesNotExist:
                        cell += 'file without grading; '
                row.append(cell)
            else:
                row.append('no file')
        ws.append(row)

    # second tab with prv data

    ws = wb.create_sheet(title='prv-aspects')

    ws['A1'] = "Students grades from {} (only Aspects related to PRV)".format(settings.NAME_PRETTY)
    ws['A1'].style = 'Headline 2'
    ws['F1'] = "Exported on: " + timestamp()
    header = ["Student id", "Student name", "Project", "Responsible teacher"]

    aspects = []  # list of PRV related aspects.
    for cat in cats:  # each category
        for aspect in cat.aspects.all():
            if 'prv' in aspect.Name.lower() or 'prv' in aspect.Description.lower():
                aspects.append([cat, aspect])
                header.append(aspect.Name)

    ws.append(header)
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
            'V', 'W', 'X', 'Y', 'Z'][:len(header)]
    for col in cols:
        ws[col + '2'].style = 'Headline 3'

    for colnr in range(1, len(header)):  # set width
        ws.column_dimensions[cols[colnr]].width = 25

    for d in des:
        reslist = []
        for c, a in aspects:
            try:
                reslist.append(d.results.get(Category=c).aspectresults.get(CategoryAspect=a).Grade)
            except:
                reslist.append('-')
        row = [d.Student.usermeta.Studentnumber, d.Student.usermeta.get_nice_fullname(), d.Proposal.Title,
               d.Proposal.ResponsibleStaff.usermeta.get_nice_name()]
        for r in reslist:
            row.append(r)
        ws.append(row)

    return wb
