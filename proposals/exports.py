#  Bep Marketplace ELE
#  Copyright (c) 2016-2022 Kolibri Solutions
#  License: See LICENSE file or https://github.com/KolibriSolutions/BepMarketplace/blob/master/LICENSE

from datetime import datetime

from django.conf import settings
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from proposals.utils import get_share_link


def get_list_projects_xlsx(proposals):
    """
    Excel export of proposals with sharelinks

    :param proposals:
    :return:
    """

    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    ws.title = "BEP Projects"

    ws['A1'] = 'Projects from {}'.format(settings.NAME_PRETTY)
    ws['A1'].style = 'Headline 2'
    ws['F1'] = "Exported on: " + str(datetime.now())

    header = ['Title', 'Track', 'Research Group', 'Responsible', 'email', 'Sharelink']
    h = ['A', 'B', 'C', 'D', 'E', 'F']

    ws.append(header)

    for hr in h:
        ws.column_dimensions[hr].width = 25
        ws[hr + '2'].style = 'Headline 3'
    ws.column_dimensions['F'].width = 100

    for p in proposals:
        ws.append([p.Title, str(p.Track), str(p.Group), p.ResponsibleStaff.usermeta.get_nice_name(), p.ResponsibleStaff.email,
                   get_share_link(p.pk)])

    return save_virtual_workbook(wb)