#  Bep Marketplace ELE
#  Copyright (c) 2016-2022 Kolibri Solutions
#  License: See LICENSE file or https://github.com/KolibriSolutions/BepMarketplace/blob/master/LICENSE
#
from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from general_view import timestamp
from django.template.defaultfilters import truncatechars


def get_list_presentations_xlsx(sets):
    """
    Lists the presentations planning with a set on each tab. Same formatting as before marketplace

    :param sets:
    :return:
    """
    wb = Workbook()
    wb.remove_sheet(wb.active)
    wss = []
    for pset in sets:
        try:
            track = pset.Track.Name
            thead = pset.Track.Head.usermeta.get_nice_name()
        except (AttributeError, ValueError):
            track = 'No track'
            thead = 'No track head'
        tit = str(pset.id) + " (" + track + ")"
        ws = wb.create_sheet(title=tit)
        wss.append(ws)
        ws['C1'] = timezone.localtime(pset.DateTime).date().strftime("%A %d %B %Y")
        ws['C3'] = "Track: " + track
        ws['C4'] = "Track responsible staff: " + thead
        ws['C5'] = "Exported on: " + timestamp()
        ws['C6'] = f'Presentation room: {pset.PresentationRoom}'
        if pset.PresentationRoom and pset.PresentationRoom.JoinLink:
            ws['C6'].hyperlink = pset.PresentationRoom.JoinLink
            ws['C6'].style = 'Hyperlink'
        ws['C7'] = f'Assessment room: {pset.AssessmentRoom}'
        if pset.AssessmentRoom and pset.AssessmentRoom.JoinLink:
            ws['C7'].hyperlink = pset.AssessmentRoom.JoinLink
            ws['C7'].style = 'Hyperlink'
        if pset.Assessors.exists():
            assessors = ''
            for a in pset.Assessors.all():
                assessors += a.usermeta.get_nice_name() + "; "
            assessors = assessors[:-2]
        else:
            assessors = '-'
        ws['C8'] = f'Assessors: {assessors}'
        if pset.PresentationAssessors.exists():
            passessors = ''
            for a in pset.PresentationAssessors.all():
                passessors += a.usermeta.get_nice_name() + "; "
            passessors = passessors[:-2]
        else:
            passessors = '-'
        ws['C9'] = f'Presentation Assessors: {passessors}'
        ws['C10'] = ''

        ws['C1'].style = 'Headline 2'
        # courses span two columns
        ws['G10'] = 'Courses'
        ws.merge_cells('G9:H9')
        ws['G10'].style = 'Headline 3'
        ws['H10'].style = 'Headline 3'
        # custom dimensions
        ws.column_dimensions['C'].width = 25  # std name
        ws.column_dimensions['E'].width = 25  # responsible name
        ws.column_dimensions['F'].width = 25  # assistants
        ws.column_dimensions['G'].width = 6  # codeBEP
        ws.column_dimensions['H'].width = 6  # codeExt
        ws.column_dimensions['I'].width = 50  # project name

        header = ["Type", "Stud. id", "Full Name", "Name", "Responsible teacher", "Assistants",
                  settings.COURSE_CODE_BEP, settings.COURSE_CODE_EXT, "Project", "Time", "Duration"]
        ws.append(header)
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws[col + '11'].style = 'Headline 3'

        for slot in pset.timeslots.all():
            if slot.CustomType:
                row = [slot.get_CustomType_display(), '', '', '', '', '', '', '', '']
            else:
                d = slot.Distribution
                row = ['', d.Student.usermeta.Studentnumber, d.Student.usermeta.get_nice_fullname(),
                       d.Student.usermeta.get_nice_name(),
                       d.Proposal.ResponsibleStaff.usermeta.get_nice_name()]
                assistants = ''
                for a in d.Proposal.Assistants.all():
                    assistants += a.usermeta.get_nice_name() + "; "
                row.append(assistants[:-2])
                row.append('x' if d.Student.usermeta.EnrolledBEP else '')
                row.append('x' if d.Student.usermeta.EnrolledExt else '')
                row.append(d.Proposal.Title)
            row.append(timezone.localtime(slot.DateTime).time().strftime("%H:%M"))
            row.append(slot.Duration())
            ws.append(row)
    return save_virtual_workbook(wb)
